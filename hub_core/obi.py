"""M-maisoku 帯替え＋許諾台帳（Wave1 V4）。

業界慣行「元付図面の帯替え」を**遵法ゲートつき**で行う（競合に無い差別化=コンプラを製品機能に）:
- 許諾台帳 = `物件/<prop>/許諾/許諾_帯替え.json`（Vaultファイル正本）。
  元付業者からの「広告掲載可・帯替え可」の許諾を、根拠ファイル（メール/書面）の
  sha256束縛つきで記録する。記録はHMAC監査に残る。
- 帯替え（obi_swap）は許諾台帳に **広告掲載 と 帯替え の両方**が無ければ拒否（fail-closed）。
- v0の加工対象は xlsx（openpyxl・元付原紙の帯行クリア＋自社帯の書込）。
  PDF帯替えは v0 対象外＝正直に拒否し、依存導入（pypdf等）は人間ゲート（tool-scout）。
"""
from __future__ import annotations

import hashlib
import json
import zipfile
from datetime import datetime, timedelta, timezone
from pathlib import Path
from xml.etree import ElementTree


def _record_hash(rec: dict) -> str:
    """許諾レコードの内容ハッシュ（監査アンカー照合用・正規化JSON）。"""
    core = {k: rec.get(k) for k in ("property", "source_company", "permitted",
                                    "evidence_sha256", "recorded_at")}
    return hashlib.sha256(json.dumps(core, ensure_ascii=False, sort_keys=True)
                          .encode("utf-8")).hexdigest()

JST = timezone(timedelta(hours=9))
PERMIT_NAME = "許諾_帯替え.json"
REQUIRED_PERMISSIONS = ("広告掲載", "帯替え")


class ObiError(Exception):
    def __init__(self, code: int, msg: str):
        super().__init__(msg)
        self.code = code
        self.msg = msg


def _now() -> str:
    return datetime.now(JST).replace(microsecond=0).isoformat()


def _prop_dir(data_dir, prop: str) -> Path:
    root = Path(data_dir).resolve()
    p = (root / "物件" / prop).resolve()
    if not p.is_relative_to(root / "物件"):   # prefix-sibling迂回封じ(R3#4)
        raise ObiError(400, f"物件名が不正: {prop!r}")
    if not p.is_dir():
        raise ObiError(404, f"物件フォルダがありません: 物件/{prop}")
    return p


def record_permission(data_dir, prop: str, source_company: str, permitted: list,
                      actor: str, evidence_rel: str = "") -> dict:
    """許諾を台帳（Vaultファイル）へ記録。
    帯替えは最高リスク許諾のため **根拠ファイル必須**（sha256束縛・F-a2是正）。
    レコード内容ハッシュ（record_hash）を返し、操作層がHMAC監査へ刻む＝ゲートの監査アンカー。"""
    pd = _prop_dir(data_dir, prop)
    if not source_company.strip():
        raise ObiError(400, "source_company（元付業者名）が必要です。")
    perms = [str(x).strip() for x in (permitted or []) if str(x).strip()]
    if not perms:
        raise ObiError(400, f"permitted（許諾内容）が必要です。例: {list(REQUIRED_PERMISSIONS)}")
    if not evidence_rel:
        raise ObiError(400, "evidence（許諾の根拠ファイル=メール/書面）は必須です。"
                            "口頭許諾は根拠メモを保存してから記録してください。")
    _root = Path(data_dir).resolve()
    evp = (_root / evidence_rel).resolve()
    if not evp.is_relative_to(_root) or not evp.is_file():   # prefix-sibling迂回封じ(R3#4)
        raise ObiError(404, f"根拠ファイルが見つかりません: {evidence_rel}")
    ev = {"evidence": evidence_rel,
          "evidence_sha256": hashlib.sha256(evp.read_bytes()).hexdigest()}
    rec = {"property": prop, "source_company": source_company.strip(),
           "permitted": perms, **ev, "recorded_by": actor, "recorded_at": _now()}
    rec["record_hash"] = _record_hash(rec)
    out = pd / "許諾"
    out.mkdir(exist_ok=True)
    (out / PERMIT_NAME).write_text(json.dumps(rec, ensure_ascii=False, indent=1), encoding="utf-8")
    return rec


def _audit_anchor(data_dir, record_hash: str) -> dict | None:
    """HMAC監査チェーンから当該許諾レコードのアンカーイベントを探す（F-a1是正）。
    **読取り時にもチェーン全体をHMAC検証**する: 偽イベントの直書き（鍵なし append）は
    チェーン破損として検出され、アンカーとして信用しない（fail-closed）。"""
    from hub_core.audit import AuditChainError, _load_chain_key, verify_audit_chain
    ap = Path(data_dir) / "audit_log.jsonl"
    if not ap.is_file():
        return None
    try:
        broken = verify_audit_chain(ap, _load_chain_key())
    except (AuditChainError, OSError, ValueError):
        raise ObiError(409, "監査チェーンの検証に失敗しました。改竄の可能性があるため出力を拒否します。")
    if broken:
        raise ObiError(409, f"監査チェーンが破損しています（seq={broken}）。"
                            "偽イベントの直書きの可能性。出力を拒否します（fail-closed）。")
    for line in ap.read_text(encoding="utf-8").splitlines():
        line = line.strip()
        if not line:
            continue
        try:
            ev = json.loads(line)
        except json.JSONDecodeError:
            continue
        if ev.get("action") == "obi_permission_recorded" and ev.get("record_hash") == record_hash:
            return ev
    return None


def check_permission(data_dir, prop: str) -> dict:
    """帯替えに必要な許諾が揃っているか。fail-closed:
    ①ファイル存在 ②必要許諾2種 ③根拠sha256一致 ④**HMAC監査チェーンにアンカー実在**
    （手書き許諾JSON=監査に無い許諾では出力できない=F-a1是正・ゲートに監査をload-bearing化）。"""
    pd = _prop_dir(data_dir, prop)
    f = pd / "許諾" / PERMIT_NAME
    if not f.is_file():
        raise ObiError(403, f"許諾記録がありません（物件/{prop}/許諾/{PERMIT_NAME}）。"
                            "元付の許諾を permission_record で記録してから出力してください。")
    rec = json.loads(f.read_text(encoding="utf-8"))
    missing = [p for p in REQUIRED_PERMISSIONS if p not in (rec.get("permitted") or [])]
    if missing:
        raise ObiError(403, f"許諾が不足: {missing}（記録済み: {rec.get('permitted')}）。出力を拒否します。")
    ev_rel = rec.get("evidence")
    if not ev_rel:
        raise ObiError(403, "根拠ファイルの無い許諾記録です。permission_record で再記録してください。")
    evp = Path(data_dir) / ev_rel
    if not evp.is_file() or hashlib.sha256(evp.read_bytes()).hexdigest() != rec.get("evidence_sha256"):
        raise ObiError(409, "許諾の根拠ファイルが変更/削除されています（sha256不一致）。再記録してください。")
    if _record_hash(rec) != rec.get("record_hash"):
        raise ObiError(409, "許諾記録の内容がハッシュと不一致（編集の痕跡）。permission_record で再記録してください。")
    anchor = _audit_anchor(data_dir, rec["record_hash"])
    if anchor is None:
        raise ObiError(403, "この許諾はHMAC監査チェーンに存在しません（正規のpermission_recordを通していない）。"
                            "出力を拒否します（fail-closed）。")
    rec["_audit_seq"] = anchor.get("seq")
    rec["_audit_hash"] = (anchor.get("entry_hash") or "")[:8]
    return rec


# 帯（取扱業者欄）は宅建業法の必要表示事項。**既定の社名を持たない**＝会社情報が未設定のまま
# 出力すると他社の広告に別会社の社名・免許番号が載るため、埋めずに fail-closed で止める
# お客様に届く面の主役は利用会社。製品名や開発元の名称を出さない。
OBI_DEFAULT = {"tel": "", "email": ""}
OBI_REQUIRED = ("company_name", "license")


def _write_excel_text(ws, row: int, value) -> None:
    """Write tenant-controlled values as OOXML strings, never formulas."""
    cell = ws.cell(row=row, column=1)
    cell.value = str(value or "")
    cell.data_type = "s"


_REJECTED_PACKAGE_PREFIXES = (
    ("xl/comments/", "コメント"),
    ("xl/threadedcomments/", "コメント"),
    ("xl/drawings/", "描画・画像"),
    ("xl/media/", "描画・画像"),
    ("xl/charts/", "描画・画像"),
    ("xl/embeddings/", "埋め込みオブジェクト"),
    ("xl/activex/", "ActiveX"),
    ("xl/ctrlprops/", "ActiveX"),
    ("xl/macrosheets/", "マクロシート"),
    ("xl/dialogsheets/", "マクロシート"),
    ("xl/externalLinks/", "外部リンク"),
    ("xl/connections.xml", "外部データ接続"),
    ("xl/querytables/", "外部データ接続"),
    ("xl/pivotcache/", "外部データ接続"),
    ("xl/datamodel/", "外部データ接続"),
    ("customui/", "カスタム実行UI"),
)


def _xml_local_name(tag: str) -> str:
    return tag.rsplit("}", 1)[-1].lower()


def _inspect_xlsx_package(path: Path) -> None:
    """Reject OOXML parts which could retain hidden PII or active content.

    Comments are rejected instead of silently deleted so an operator cannot mistake a
    workbook with review notes for a fully reviewed source. Non-target worksheets are
    handled separately: they are accepted as input but never copied to the output.
    """
    try:
        with zipfile.ZipFile(path) as archive:
            names = archive.namelist()
            normalized_names = [name.replace("\\", "/").lower() for name in names]
            if len(normalized_names) != len(set(normalized_names)):
                raise ObiError(400, "元図面のパッケージに重複した部品名があります。")
            lower_names = {name.lower(): name for name in names}
            if "[content_types].xml" not in lower_names or "xl/workbook.xml" not in lower_names:
                raise ObiError(400, "有効な .xlsx パッケージではありません。")

            for prefix, label in _REJECTED_PACKAGE_PREFIXES:
                if any(name.lower().startswith(prefix.lower()) for name in names):
                    raise ObiError(400, f"元図面に{label}が含まれるため帯替えを拒否します。")
            if any(name.lower().endswith("vbaproject.bin") for name in names):
                raise ObiError(400, "元図面にマクロが含まれるため帯替えを拒否します。")

            xml_members = [
                name for name in names
                if name.lower().startswith("xl/worksheets/") and name.lower().endswith(".xml")
            ]
            relation_members = [name for name in names if name.lower().endswith(".rels")]
            for name in xml_members + relation_members:
                try:
                    root = ElementTree.fromstring(archive.read(name))
                except (ElementTree.ParseError, KeyError, OSError) as exc:
                    raise ObiError(400, f"元図面のXMLが不正です: {name}") from exc
                for element in root.iter():
                    local_name = _xml_local_name(element.tag)
                    if name in xml_members and local_name == "f":
                        raise ObiError(400, "元図面に数式が含まれるため帯替えを拒否します。")
                    if name in xml_members and local_name == "hyperlink":
                        raise ObiError(400, "元図面にハイパーリンクが含まれるため帯替えを拒否します。")
                    if name in relation_members:
                        target_mode = str(element.attrib.get("TargetMode") or "").lower()
                        rel_type = str(element.attrib.get("Type") or "").lower()
                        if target_mode == "external" or rel_type.endswith("/externallink"):
                            raise ObiError(400, "元図面に外部リンクが含まれるため帯替えを拒否します。")
    except zipfile.BadZipFile as exc:
        raise ObiError(400, "有効な .xlsx パッケージではありません。") from exc


def _inspect_loaded_workbook(wb) -> None:
    """Second-line inspection for content normalized by openpyxl."""
    for sheet in wb.worksheets:
        if getattr(sheet, "_images", None) or getattr(sheet, "_charts", None):
            raise ObiError(400, "元図面に描画・画像が含まれるため帯替えを拒否します。")
        # Iterating actual populated cells avoids expanding a hostile max_row/max_column range.
        for cell in sheet._cells.values():
            if cell.data_type == "f":
                raise ObiError(400, "元図面に数式が含まれるため帯替えを拒否します。")
            if cell.hyperlink is not None:
                raise ObiError(400, "元図面にハイパーリンクが含まれるため帯替えを拒否します。")
            if cell.comment is not None:
                raise ObiError(400, "元図面にコメントが含まれるため帯替えを拒否します。")


def _sanitize_workbook(wb):
    """Keep only the public target sheet and remove source-author metadata."""
    target = wb.active
    if target is None or target.sheet_state != "visible":
        raise ObiError(400, "公開対象の表示シートを特定できません。")
    _inspect_loaded_workbook(wb)
    for sheet in list(wb.worksheets):
        if sheet is not target:
            wb.remove(sheet)
    target.title = "マイソク"
    target.sheet_properties.codeName = None
    wb.code_name = None
    wb._external_links = []
    wb.defined_names.clear()
    if hasattr(wb, "custom_doc_props"):
        wb.custom_doc_props.props.clear()
    for field in (
        "creator", "title", "description", "subject", "identifier", "language",
        "lastModifiedBy", "category", "contentStatus", "version", "revision", "keywords",
    ):
        setattr(wb.properties, field, None)
    return target


def _validate_sanitized_output(path: Path) -> None:
    """Fail closed if serialization reintroduced anything outside the public sheet."""
    _inspect_xlsx_package(path)
    try:
        with zipfile.ZipFile(path) as archive:
            worksheets = [
                name for name in archive.namelist()
                if name.lower().startswith("xl/worksheets/") and name.lower().endswith(".xml")
            ]
    except zipfile.BadZipFile as exc:
        raise ObiError(500, "帯替え出力の再検証に失敗しました。") from exc
    if len(worksheets) != 1:
        raise ObiError(500, "帯替え出力に公開対象外のシートデータが残っています。")
    import openpyxl
    try:
        wb = openpyxl.load_workbook(path, data_only=False, keep_links=False)
    except (OSError, ValueError, KeyError) as exc:
        raise ObiError(500, "帯替え出力の再検証に失敗しました。") from exc
    try:
        if len(wb.worksheets) != 1 or wb.active is None or wb.active.sheet_state != "visible":
            raise ObiError(500, "帯替え出力に公開対象外のシートが残っています。")
        _inspect_loaded_workbook(wb)
    finally:
        wb.close()


def swap_obi_xlsx(data_dir, prop: str, source_rel: str, actor: str,
                  company: dict | None = None, clear_rows: str = "") -> dict:
    """xlsx 元付図面の帯替え。

    数式・リンク・コメント・描画等は fail-closed で拒否する。受理した原紙も、公開対象の
    アクティブシート1枚だけを出力し、hiddenを含む他シートは再配布しない。
    """
    permit = check_permission(data_dir, prop)
    root = Path(data_dir).resolve()
    src = (root / source_rel).resolve()
    if not src.is_relative_to(root) or not src.is_file():   # prefix-sibling迂回封じ(R3#4)
        raise ObiError(404, f"元図面が見つかりません: {source_rel}")
    if src.suffix.lower() != ".xlsx":
        raise ObiError(400, f"v0の帯替え対象は .xlsx のみ（{src.suffix} は未対応=依存導入は人間ゲート）。")
    clear_rows = str(clear_rows or "").strip()
    if not clear_rows:
        raise ObiError(
            400,
            "元の取扱業者欄を削除する行範囲（clear_rows）が必要です。"
            "原紙を確認し、'開始-終了' 形式で指定してください。",
        )
    try:
        a, b = clear_rows.split("-", 1)
        r0, r1 = int(a), int(b)
        assert 1 <= r0 <= r1 and (r1 - r0 + 1) <= 500
    except (ValueError, AssertionError):
        raise ObiError(
            400,
            f"clear_rows は500行以内の '開始-終了' 形式の行番号: {clear_rows!r}",
        )
    rows_to_clear = (r0, r1)
    c = {**OBI_DEFAULT, **(company or {})}
    missing = [k for k in OBI_REQUIRED if not str(c.get(k) or "").strip()]
    if missing:
        raise ObiError(400, "取扱業者欄に載せる会社情報が未設定です（"
                            + "・".join(missing)
                            + "）。「業者情報」の画面で自社の社名と免許番号を登録してから"
                              "出力してください。別会社の名前を代わりに載せることはしません。")
    import openpyxl
    _inspect_xlsx_package(src)
    try:
        wb = openpyxl.load_workbook(src, data_only=False, keep_links=False)
    except (OSError, ValueError, KeyError) as exc:
        raise ObiError(400, "元図面を安全に読み込めませんでした。") from exc
    try:
        ws = _sanitize_workbook(wb)
    except Exception:
        wb.close()
        raise
    r0, r1 = rows_to_clear
    if r1 > ws.max_row:
        wb.close()
        raise ObiError(
            400,
            f"clear_rows {clear_rows!r} が原紙の最終行 {ws.max_row} を超えています。",
        )
    # 値だけ消すと旧社色・罫線・結合セルが残り、空の旧帯として印刷される。
    # 交差する結合を解いたうえで行そのものを削除し、旧帯の見た目も残さない。
    for merged in list(ws.merged_cells.ranges):
        if merged.min_row <= r1 and merged.max_row >= r0:
            ws.unmerge_cells(str(merged))
    ws.delete_rows(r0, r1 - r0 + 1)
    base = ws.max_row + 2
    _write_excel_text(ws, base, "【取扱】" + c["company_name"])
    _write_excel_text(ws, base + 1, c["license"])
    contact = "・".join(v for v in (c.get("tel"), c.get("email")) if v)
    if contact:
        _write_excel_text(ws, base + 2, contact)
    _write_excel_text(
        ws,
        base + 3,
        f"元付: {permit['source_company']}（帯替え許諾 {permit['recorded_at'][:10]} 記録・"
        f"監査#{permit.get('_audit_seq', '?')}/{permit.get('_audit_hash', '')}）",
    )
    pd = _prop_dir(data_dir, prop)
    outdir = pd / "書類" / "マイソク帯替え"
    outdir.mkdir(parents=True, exist_ok=True)
    ver = 1 + sum(1 for x in outdir.glob("v*") if x.is_dir())
    vdir = outdir / f"v{ver}"
    vdir.mkdir()
    out = vdir / f"マイソク帯替え_{prop}.xlsx"
    try:
        wb.save(out)
        _validate_sanitized_output(out)
    except Exception:
        out.unlink(missing_ok=True)
        try:
            vdir.rmdir()
        except OSError:
            pass
        raise
    finally:
        wb.close()
    return {"out": str(out.relative_to(root)), "version": ver,
            "source_company": permit["source_company"]}
