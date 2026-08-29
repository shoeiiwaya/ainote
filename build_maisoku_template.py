#!/usr/bin/env python3
"""Build the three original, redistributable Excel maisoku templates from blank workbooks.

The deleted predecessor was a third-party business workbook. This builder does not
read or transform any existing Office file. A is the default A4 landscape output,
B is an A4 portrait buyer handout, and C is a monochrome A4 landscape FAX sheet.
Every workbook contains one visible sheet, named ranges for the fill contract,
image anchors, and provenance in document properties.
"""
from __future__ import annotations

import argparse
import hashlib
import re
import tempfile
import zipfile
from copy import copy
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import Workbook
from openpyxl.packaging.custom import StringProperty
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils.cell import absolute_coordinate, get_column_letter
from openpyxl.workbook.defined_name import DefinedName


ROOT = Path(__file__).resolve().parent
DEFAULT_OUTPUT = ROOT / "hub_core" / "templates" / "office" / "maisoku_genshi.xlsx"
TEMPLATE_OUTPUTS = {
    "A": ROOT / "hub_core" / "templates" / "office" / "maisoku_a_standard_landscape.xlsx",
    "B": ROOT / "hub_core" / "templates" / "office" / "maisoku_b_buyer_portrait.xlsx",
    "C": ROOT / "hub_core" / "templates" / "office" / "maisoku_c_fax_mono.xlsx",
}
SHEET_NAME = "マイソク原紙"
TEMPLATE_IDS = {
    "A": "ainote-maisoku-a-standard-landscape-v1",
    "B": "ainote-maisoku-b-buyer-portrait-v1",
    "C": "ainote-maisoku-c-fax-mono-v1",
}
TEMPLATE_ID = TEMPLATE_IDS["A"]
FIXED_TIME = datetime(2026, 8, 7, tzinfo=timezone.utc)

INK = "202328"
MUTED = "5F6770"
RULE = "C9CED4"
LABEL = "F0F2F4"
SOFT = "E1E5E9"
ACCENT = "2A2E37"
WHITE = "FFFFFF"


def _rgb_dark(value) -> bool:
    rgb = str(value or "")[-6:]
    if not re.fullmatch(r"[0-9A-Fa-f]{6}", rgb):
        return False
    red, green, blue = (int(rgb[index:index + 2], 16) for index in (0, 2, 4))
    return (red * 299 + green * 587 + blue * 114) < 150000


FIELD_CELLS = {
    "property_type": "A1",
    "property_name": "E1",
    "title_copy": "A3",
    "lead": "I3",
    "price_label": "A4",
    "price": "E4",
    "price_note": "I39",
    "yield_rate": "N4",
    "income_basis_label": "O39",
    "owner_change": "S4",
    "address": "D9",
    "access": "N9",
    "land_area": "D10",
    "building_area": "N10",
    "floor_plan": "D11",
    "built": "N11",
    "structure": "D12",
    "total_units": "N12",
    "genkyo": "D13",
    "hikiwatashi": "N13",
    "nearest_station": "D14",
    "torihiki_taiyo": "N14",
    "floors_total": "B7",
    "balcony_area": "H7",
    "chikunensu": "M7",
    "bunjou_company": "Q7",
    "sekou_company": "V7",
    "land_right": "D16",
    "chimoku": "N16",
    "toshi_keikaku": "D17",
    "youto": "N17",
    "kenpei_yoseki": "D18",
    "bouka": "N18",
    "other_law": "D19",
    "road": "N19",
    "shidou": "D20",
    "direction": "N20",
    "management_fee": "D22",
    "repair_fund": "N22",
    "other_fee": "D23",
    "parking": "N23",
    "monthly_rent": "D24",
    "annual_income": "N24",
    "lease_period": "D25",
    "kanri_company": "N25",
    "sales_points": "D27",
    "route_lines": "D28",
    "surroundings": "D29",
    "setsubi": "D30",
    "photo_main_tag": "A32",
    "photo_main_caption": "F32",
    "photo_sub1_tag": "A33",
    "photo_sub1_caption": "F33",
    "photo_sub2_tag": "A34",
    "photo_sub2_caption": "F34",
    "photo_sub3_tag": "M32",
    "photo_sub3_caption": "R32",
    "photo_map_tag": "M33",
    "photo_map_caption": "R33",
    "photo_floorplan_caption": "R34",
    "reins_no": "D35",
    "ad_permission": "J35",
    "obi_swap": "P35",
    "key_handling": "V35",
    "bikou": "D36",
    "yuko_kigen": "D38",
    "published": "L38",
    "next_update": "T38",
    "staging_note": "C39",
    "special_notes": "U39",
    "company_name": "A41",
    "license": "Q41",
    "association": "Q42",
    "company_address": "D43",
    "company_tel": "O43",
    "company_fax": "U43",
    "company_email": "D44",
    "fair_trade_association": "Q44",
    "staff": "D45",
    "holiday": "K45",
    "fee": "Q45",
}

STYLE_RANGES = {
    "style_font_body": ["A1:X45"],
    "style_font_title": ["A1:X6", "A8:X8", "A15:X15", "A21:X21", "A26:X26",
                         "A31:X31", "A37:X37", "A40:X40", "A41:L42"],
    "style_accent_fill": ["A1:D2", "A4:D6", "A8:X8", "A15:X15", "A21:X21",
                           "A26:X26", "A31:X31", "A37:X37", "A40:X40"],
    "style_accent_ink": ["A1:D2", "A4:D6", "A8:X8", "A15:X15", "A21:X21",
                          "A26:X26", "A31:X31", "A37:X37", "A40:X40"],
    "style_accent_text": ["A3:X3", "K4:X6", "A41:L42"],
    "style_accent_soft": ["K4:X6"],
}


def _merge_value(ws, cell_range: str, value: str = "") -> None:
    ws.merge_cells(cell_range)
    ws[cell_range.split(":", 1)[0]] = value


def _style_range(ws, cell_range: str, *, fill: str | None = None,
                 font: Font | None = None, alignment: Alignment | None = None,
                 border: Border | None = None) -> None:
    for row in ws[cell_range]:
        for cell in row:
            if fill:
                cell.fill = PatternFill("solid", fgColor=fill)
            if font:
                cell.font = font
            if alignment:
                cell.alignment = alignment
            if border:
                cell.border = border


def _define(wb: Workbook, name: str, target: str) -> None:
    wb.defined_names.add(
        DefinedName(name, attr_text=f"'{SHEET_NAME}'!{absolute_coordinate(target)}")
    )


def _define_range(wb: Workbook, name: str, target: str) -> None:
    start, end = target.split(":", 1)
    wb.defined_names.add(
        DefinedName(
            name,
            attr_text=(f"'{SHEET_NAME}'!{absolute_coordinate(start)}:"
                       f"{absolute_coordinate(end)}"),
        )
    )


def build_template(out_path: Path = DEFAULT_OUTPUT, *, variant: str = "A") -> Path:
    variant = str(variant or "A").upper()
    if variant not in TEMPLATE_IDS:
        raise ValueError("variant must be A, B, or C")
    wb = Workbook()
    ws = wb.active
    ws.title = SHEET_NAME
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.orientation = "portrait" if variant == "B" else "landscape"
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.print_area = "A1:X45"
    ws.sheet_properties.outlinePr.summaryBelow = False
    ws.freeze_panes = None
    ws.page_margins.left = 0.25
    ws.page_margins.right = 0.25
    ws.page_margins.top = 0.25
    ws.page_margins.bottom = 0.25
    ws.page_margins.footer = 0.1
    ws.oddFooter.center.text = "&P / &N"
    ws.oddFooter.center.size = 7
    ws.oddFooter.center.color = MUTED

    for col in range(1, 25):
        ws.column_dimensions[ws.cell(1, col).column_letter].width = 7.0
    for row in range(1, 46):
        ws.row_dimensions[row].height = 10.5
    for row, height in {1: 18, 2: 18, 3: 18, 4: 18, 5: 18, 6: 18, 7: 14,
                        8: 14, 15: 14, 21: 14, 26: 14, 31: 14, 37: 14, 40: 14,
                        27: 34, 28: 16, 29: 16, 30: 16,
                        32: 26, 33: 26, 34: 26, 35: 14, 36: 34,
                        39: 20, 41: 16, 42: 16}.items():
        ws.row_dimensions[row].height = height

    # Header and primary commercial facts.
    _merge_value(ws, "A1:D2", "物件種目")
    _merge_value(ws, "E1:X2")
    _merge_value(ws, "A3:H3")
    _merge_value(ws, "I3:X3")
    _merge_value(ws, "A4:D6", "価格")
    _merge_value(ws, "E4:J6")
    _merge_value(ws, "K4:M6", "想定利回り")
    _merge_value(ws, "N4:R6")
    _merge_value(ws, "S4:X6")

    for label_range, value_range, label in (
        ("A7:A7", "B7:E7", "階建"),
        ("F7:G7", "H7:J7", "バルコニー"),
        ("K7:L7", "M7:O7", "築年数"),
        ("P7:P7", "Q7:T7", "分譲"),
        ("U7:U7", "V7:X7", "施工"),
    ):
        _merge_value(ws, label_range, label)
        _merge_value(ws, value_range)

    # Two-column fact sections.
    sections = {
        8: "物件概要", 15: "法令・権利", 21: "費用・収益",
        26: "特徴・設備", 31: "写真・間取り", 37: "掲載情報", 40: "取扱業者",
    }
    for row, title in sections.items():
        _merge_value(ws, f"A{row}:X{row}", title)

    facts = [
        (9, "所在地", "交通"), (10, "土地・敷地面積", "建物・専有面積"),
        (11, "間取り", "築年月"), (12, "構造", "総戸数"),
        (13, "現況", "引渡し"), (14, "最寄駅", "取引態様"),
        (16, "土地権利", "地目"), (17, "都市計画", "用途地域"),
        (18, "建ぺい率・容積率", "防火規制"), (19, "他の法令上の制限", "接道状況"),
        (20, "私道負担", "方位"), (22, "管理費", "修繕積立金"),
        (23, "その他費用", "駐車場・駐輪場"), (24, "月額賃料・満室想定", "年間収入・満室想定"),
        (25, "賃貸借期間", "管理会社"),
    ]
    for row, left_label, right_label in facts:
        _merge_value(ws, f"A{row}:C{row}", left_label)
        _merge_value(ws, f"D{row}:J{row}")
        _merge_value(ws, f"K{row}:M{row}", right_label)
        _merge_value(ws, f"N{row}:X{row}")

    for row, label in ((27, "セールスポイント"), (28, "沿線"),
                       (29, "周辺環境"), (30, "設備"), (36, "備考")):
        _merge_value(ws, f"A{row}:C{row}", label)
        _merge_value(ws, f"D{row}:X{row}")

    for label_range, value_range, label in (
        ("A32:E32", "F32:L32", "写真1"),
        ("A33:E33", "F33:L33", "写真2"),
        ("A34:E34", "F34:L34", "写真3"),
        ("M32:Q32", "R32:X32", "写真4"),
        ("M33:Q33", "R33:X33", "地図"),
        ("M34:Q34", "R34:X34", "間取り図"),
    ):
        _merge_value(ws, label_range, label)
        _merge_value(ws, value_range)
    for label_range, value_range, label in (
        ("A35:C35", "D35:F35", "REINS"),
        ("G35:I35", "J35:L35", "広告"),
        ("M35:O35", "P35:R35", "帯替え"),
        ("S35:U35", "V35:X35", "鍵"),
    ):
        _merge_value(ws, label_range, label)
        _merge_value(ws, value_range)

    _merge_value(ws, "A38:C38", "有効期限")
    _merge_value(ws, "D38:H38")
    _merge_value(ws, "I38:K38", "情報公開日")
    _merge_value(ws, "L38:P38")
    _merge_value(ws, "Q38:S38", "次回更新予定日")
    _merge_value(ws, "T38:X38")
    for label_range, value_range, label in (
        ("A39:B39", "C39:F39", "下書き"),
        ("G39:H39", "I39:L39", "価格注記"),
        ("M39:N39", "O39:R39", "収益根拠"),
        ("S39:T39", "U39:X39", "特記"),
    ):
        _merge_value(ws, label_range, label)
        _merge_value(ws, value_range)

    _merge_value(ws, "A41:L42")
    _merge_value(ws, "M41:P41", "免許番号")
    _merge_value(ws, "Q41:X41")
    _merge_value(ws, "M42:P42", "保証協会")
    _merge_value(ws, "Q42:X42")
    _merge_value(ws, "A43:C43", "所在地")
    _merge_value(ws, "D43:L43")
    _merge_value(ws, "M43:N43", "TEL")
    _merge_value(ws, "O43:R43")
    _merge_value(ws, "S43:T43", "FAX")
    _merge_value(ws, "U43:X43")
    _merge_value(ws, "A44:C44", "Email")
    _merge_value(ws, "D44:L44")
    _merge_value(ws, "M44:P44", "公取協")
    _merge_value(ws, "Q44:X44")
    _merge_value(ws, "A45:C45", "担当")
    _merge_value(ws, "D45:H45")
    _merge_value(ws, "I45:J45", "定休")
    _merge_value(ws, "K45:N45")
    _merge_value(ws, "O45:P45", "手数料")
    _merge_value(ws, "Q45:X45")

    thin = Side(style="thin", color=RULE)
    medium = Side(style="medium", color=ACCENT)
    grid = Border(left=thin, right=thin, top=thin, bottom=thin)
    outer = Border(left=medium, right=medium, top=medium, bottom=medium)
    body_font = Font(name="Yu Gothic", size=8.5, color=INK)
    label_font = Font(name="Yu Gothic", size=7.5, bold=True, color=MUTED)
    section_font = Font(name="Yu Gothic", size=9, bold=True, color=WHITE)
    centered = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)

    _style_range(ws, "A1:X45", font=body_font, alignment=left)
    _style_range(ws, "A1:X45", border=grid)
    for row in sections:
        _style_range(ws, f"A{row}:X{row}", fill=ACCENT, font=section_font, alignment=left)
    for row, _l, _r in facts:
        _style_range(ws, f"A{row}:C{row}", fill=LABEL, font=label_font, alignment=centered)
        _style_range(ws, f"K{row}:M{row}", fill=LABEL, font=label_font, alignment=centered)
    for cell_range in ("A7:A7", "F7:G7", "K7:L7", "P7:P7", "U7:U7",
                       "A35:C35", "G35:I35", "M35:O35", "S35:U35"):
        _style_range(ws, cell_range, fill=LABEL, font=label_font, alignment=centered)
    for cell_range in ("B7:E7", "H7:J7", "M7:O7", "Q7:T7", "V7:X7",
                       "D35:F35", "J35:L35", "P35:R35", "V35:X35"):
        _style_range(
            ws, cell_range,
            alignment=Alignment(horizontal="left", vertical="center", shrink_to_fit=True),
        )
    for row in (27, 28, 29, 30, 36):
        _style_range(ws, f"A{row}:C{row}", fill=LABEL, font=label_font, alignment=centered)
    for cell_range in ("A38:C38", "I38:K38", "Q38:S38", "M41:P41", "M42:P42",
                       "A43:C43", "M43:N43", "S43:T43", "A44:C44", "M44:P44",
                       "A45:C45", "I45:J45", "O45:P45", "A39:B39", "G39:H39",
                       "M39:N39", "S39:T39", "A32:E32", "A33:E33", "A34:E34",
                       "M32:Q32", "M33:Q33", "M34:Q34"):
        _style_range(ws, cell_range, fill=LABEL, font=label_font, alignment=centered)

    _style_range(ws, "A1:D2", fill=ACCENT,
                 font=Font(name="Yu Gothic", size=12, bold=True, color=WHITE), alignment=centered)
    _style_range(ws, "E1:X2", font=Font(name="Yu Gothic", size=21, bold=True, color=INK), alignment=left)
    _style_range(ws, "A3:H3", font=Font(name="Yu Gothic", size=10.5, bold=True, color=ACCENT), alignment=left)
    _style_range(ws, "I3:X3", font=Font(name="Yu Gothic", size=8.5, color=ACCENT), alignment=left)
    _style_range(ws, "A4:D6", fill=ACCENT,
                 font=Font(name="Yu Gothic", size=10, bold=True, color=WHITE), alignment=centered)
    _style_range(ws, "E4:J6", font=Font(name="Yu Gothic", size=22, bold=True, color=INK), alignment=centered)
    _style_range(ws, "K4:X6", fill=SOFT, alignment=centered)
    _style_range(ws, "K4:M6", font=Font(name="Yu Gothic", size=8, bold=True, color=ACCENT), alignment=centered)
    _style_range(ws, "N4:R6", font=Font(name="Yu Gothic", size=16, bold=True, color=ACCENT), alignment=centered)
    _style_range(ws, "S4:X6", font=Font(name="Yu Gothic", size=9, bold=True, color=INK), alignment=centered)
    for cell_range in ("F32:L32", "F33:L33", "F34:L34", "R32:X32", "R33:X33", "R34:X34",
                       "C39:F39", "I39:L39", "O39:R39", "U39:X39"):
        _style_range(
            ws, cell_range,
            alignment=Alignment(horizontal="left", vertical="center", shrink_to_fit=True),
        )
    _style_range(ws, "A41:L42", font=Font(name="Yu Gothic", size=15, bold=True, color=ACCENT), alignment=left)

    # Rows 32-36 are the two visual anchors. The empty originals contain no media;
    # validated property bytes are embedded by docgen at export time.
    _define(wb, "image_photo_main", "A32")
    _define(wb, "image_floorplan", "M32")

    if variant == "B":
        # Buyer sheet: portrait and slightly narrower columns. The same deterministic
        # named-range contract is retained so no field is silently dropped.
        for col in range(1, 25):
            ws.column_dimensions[get_column_letter(col)].width = 5.25
        ws.page_margins.left = 0.2
        ws.page_margins.right = 0.2
    elif variant == "C":
        # Broker FAX: remove tenant colour from the original itself. docgen reapplies
        # monochrome after field/theme filling as a second line of defence.
        for row in ws.iter_rows(min_row=1, max_row=45, min_col=1, max_col=24):
            for cell in row:
                if cell.fill.fill_type == "solid":
                    cell.fill = PatternFill("solid", fgColor=("222222" if _rgb_dark(cell.fill.fgColor.rgb) else "E6E6E6"))
                font = copy(cell.font)
                font.color = "FFFFFF" if cell.fill.fill_type == "solid" and _rgb_dark(cell.fill.fgColor.rgb) else "111111"
                cell.font = font

    for key, cell in FIELD_CELLS.items():
        _define(wb, f"ms_{key}", cell)
    for role, ranges in STYLE_RANGES.items():
        for index, cell_range in enumerate(ranges, 1):
            _define_range(wb, f"{role}_{index:02d}", cell_range)

    wb.properties.creator = "Original workbook generator"
    wb.properties.lastModifiedBy = "Original workbook generator"
    wb.properties.title = {
        "A": "マイソク A 標準A4横",
        "B": "マイソク B 買主向けA4縦",
        "C": "マイソク C 業者間FAX白黒",
    }[variant]
    wb.properties.subject = "不動産販売図面"
    wb.properties.description = (
        "Created from a blank workbook by build_maisoku_template.py; "
        "contains no third-party template content."
    )
    wb.properties.keywords = "maisoku,original-template"
    wb.properties.created = FIXED_TIME
    wb.properties.modified = FIXED_TIME
    wb.custom_doc_props.append(StringProperty(name="TemplateId", value=TEMPLATE_IDS[variant]))
    wb.custom_doc_props.append(StringProperty(name="MaisokuVariant", value=variant))
    wb.custom_doc_props.append(StringProperty(
        name="Provenance", value="Created from a blank workbook; no third-party template used"))
    wb.custom_doc_props.append(StringProperty(name="SchemaVersion", value="1"))

    out_path = Path(out_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    return out_path


def normalized_package_digest(path: Path) -> str:
    """Hash OOXML member names and bytes; ZIP timestamps do not affect this digest."""
    h = hashlib.sha256()
    with zipfile.ZipFile(path) as zf:
        for name in sorted(zf.namelist()):
            payload = zf.read(name)
            if name == "docProps/core.xml":
                # openpyxl replaces dcterms:modified at save time.
                payload = re.sub(
                    rb"(<dcterms:modified[^>]*>)[^<]*(</dcterms:modified>)",
                    rb"\g<1>GENERATED\g<2>", payload,
                )
            h.update(name.encode("utf-8"))
            h.update(b"\0")
            h.update(payload)
            h.update(b"\0")
    return h.hexdigest()


def _variant_for_path(path: Path) -> str:
    resolved = Path(path).resolve()
    for variant, candidate in TEMPLATE_OUTPUTS.items():
        if resolved == candidate.resolve():
            return variant
    return "A"


def check_template(path: Path = DEFAULT_OUTPUT, *, variant: str | None = None) -> bool:
    if not path.is_file():
        return False
    with tempfile.TemporaryDirectory(prefix="maisoku_template_check_") as td:
        rebuilt = build_template(Path(td) / path.name,
                                 variant=(variant or _variant_for_path(path)))
        return normalized_package_digest(path) == normalized_package_digest(rebuilt)


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("output", nargs="?", type=Path, default=DEFAULT_OUTPUT)
    parser.add_argument("--variant", choices=("A", "B", "C"), default="A")
    parser.add_argument("--all", action="store_true",
                        help="build all three canonical A/B/C templates")
    parser.add_argument("--check", action="store_true",
                        help="verify that the committed workbook matches this builder")
    args = parser.parse_args()
    if args.all:
        for variant, output in TEMPLATE_OUTPUTS.items():
            print(build_template(output, variant=variant))
        return 0
    if args.check:
        ok = check_template(args.output, variant=args.variant)
        print("template matches builder" if ok else "template differs from builder")
        return 0 if ok else 1
    print(build_template(args.output, variant=args.variant))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
